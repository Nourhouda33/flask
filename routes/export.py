"""
Blueprint export — Export PDF et CSV/Excel des données médicales.

Endpoints :
  GET  /api/export/consultation/<id>/pdf  → PDF du rapport médical (ReportLab)
  POST /api/export/query-results          → Export CSV ou Excel des résultats SQL
"""

import io
import csv
import time
import logging
from datetime import datetime

from flask import Blueprint, request, g, send_file, current_app

from database.db import db
from models.consultation import Consultation
from auth.decorators import token_required, role_required
from utils.response_helper import error_response

logger    = logging.getLogger(__name__)
export_bp = Blueprint("export", __name__)

# ─────────────────────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _check_sql_blacklist(sql: str) -> bool:
    """Vérifie que le SQL ne contient pas d'opérations dangereuses."""
    import re
    dangerous = [r"\bDROP\b", r"\bTRUNCATE\b", r"\bALTER\b", r"\bCREATE\b",
                 r"\bGRANT\b", r"\bREVOKE\b", r"\bINTO\s+OUTFILE\b"]
    upper = sql.upper()
    return not any(re.search(p, upper) for p in dangerous)


# ─────────────────────────────────────────────────────────────────────────────
#  GET /api/export/consultation/<id>/pdf
# ─────────────────────────────────────────────────────────────────────────────

@export_bp.route("/consultation/<int:consultation_id>/pdf", methods=["GET"])
@token_required
@role_required("admin", "doctor", "staff")
def export_consultation_pdf(consultation_id: int):
    """
    Génère un rapport PDF pour une consultation médicale via ReportLab.

    Returns:
        200 : Fichier PDF (application/pdf)
        404 : Consultation introuvable.
        500 : Erreur de génération PDF.
    """
    consultation = Consultation.query.get(consultation_id)
    if not consultation:
        return error_response(f"Consultation {consultation_id} introuvable", 404)

    # Charger les relations
    patient = consultation.patient
    staff   = consultation.staff

    try:
        pdf_buffer = _generate_consultation_pdf(consultation, patient, staff)
    except ImportError:
        return error_response(
            "ReportLab n'est pas installé. Installez-le avec : pip install reportlab",
            500,
            error="REPORTLAB_NOT_INSTALLED",
        )
    except Exception as exc:
        logger.error("Erreur génération PDF consultation %s : %s", consultation_id, str(exc), exc_info=True)
        return error_response(f"Erreur lors de la génération du PDF : {str(exc)[:200]}", 500)

    filename = f"consultation_{consultation_id}_{datetime.utcnow().strftime('%Y%m%d')}.pdf"

    logger.info(
        "PDF généré — consultation_id=%s par user_id=%s",
        consultation_id, g.current_user.id_user,
    )

    return send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )


def _generate_consultation_pdf(consultation, patient, staff) -> io.BytesIO:
    """
    Génère le PDF d'un rapport de consultation avec ReportLab.

    Args:
        consultation: Instance Consultation.
        patient:      Instance Patient (peut être None).
        staff:        Instance Medical_staff (peut être None).

    Returns:
        Buffer BytesIO contenant le PDF.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    )

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()

    # Styles personnalisés
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Title"],
        fontSize=18,
        textColor=colors.HexColor("#1a5276"),
        spaceAfter=6,
    )
    heading_style = ParagraphStyle(
        "CustomHeading",
        parent=styles["Heading2"],
        fontSize=12,
        textColor=colors.HexColor("#2874a6"),
        spaceBefore=12,
        spaceAfter=4,
    )
    body_style = ParagraphStyle(
        "CustomBody",
        parent=styles["Normal"],
        fontSize=10,
        leading=14,
    )
    label_style = ParagraphStyle(
        "Label",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#666666"),
    )

    story = []

    # ── En-tête ────────────────────────────────────────────────────────────
    story.append(Paragraph("Healthcare AI Platform", title_style))
    story.append(Paragraph("Rapport de Consultation Médicale", styles["Heading1"]))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#1a5276")))
    story.append(Spacer(1, 0.5 * cm))

    # ── Informations de la consultation ────────────────────────────────────
    story.append(Paragraph("Informations Générales", heading_style))

    consult_date = consultation.date.strftime("%d/%m/%Y à %H:%M") if consultation.date else "Non renseignée"
    info_data = [
        ["N° Consultation :", str(consultation.id_consultation)],
        ["Date :",            consult_date],
        ["Médecin :",         staff.name_staff if staff else "Non renseigné"],
        ["Spécialité :",      staff.speciality if staff and staff.speciality else "—"],
    ]
    info_table = Table(info_data, colWidths=[4 * cm, 12 * cm])
    info_table.setStyle(TableStyle([
        ("FONTNAME",    (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE",    (0, 0), (-1, -1), 10),
        ("FONTNAME",    (0, 0), (0, -1), "Helvetica-Bold"),
        ("TEXTCOLOR",   (0, 0), (0, -1), colors.HexColor("#2874a6")),
        ("VALIGN",      (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#dee2e6")),
        ("PADDING",     (0, 0), (-1, -1), 6),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.5 * cm))

    # ── Informations patient ───────────────────────────────────────────────
    if patient:
        story.append(Paragraph("Patient", heading_style))
        patient_name = f"{patient.first_name} {patient.last_name}"
        patient_data = [
            ["Nom complet :", patient_name],
            ["Date de naissance :", patient.birthdate.strftime("%d/%m/%Y") if patient.birthdate else "—"],
            ["Âge :",         f"{patient.age} ans" if patient.age else "—"],
            ["Genre :",       "Masculin" if patient.gender == "Male" else "Féminin" if patient.gender == "Female" else "—"],
        ]
        patient_table = Table(patient_data, colWidths=[4 * cm, 12 * cm])
        patient_table.setStyle(TableStyle([
            ("FONTNAME",    (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE",    (0, 0), (-1, -1), 10),
            ("FONTNAME",    (0, 0), (0, -1), "Helvetica-Bold"),
            ("TEXTCOLOR",   (0, 0), (0, -1), colors.HexColor("#2874a6")),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
            ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#dee2e6")),
            ("PADDING",     (0, 0), (-1, -1), 6),
        ]))
        story.append(patient_table)
        story.append(Spacer(1, 0.5 * cm))

    # ── Diagnostic ────────────────────────────────────────────────────────
    story.append(Paragraph("Diagnostic", heading_style))
    story.append(Paragraph(consultation.diagnosis or "Non renseigné", body_style))
    story.append(Spacer(1, 0.3 * cm))

    # ── Traitement ────────────────────────────────────────────────────────
    story.append(Paragraph("Traitement Prescrit", heading_style))
    story.append(Paragraph(consultation.treatment or "Aucun traitement prescrit", body_style))
    story.append(Spacer(1, 0.3 * cm))

    # ── Rapport médical ───────────────────────────────────────────────────
    if consultation.medical_report:
        story.append(Paragraph("Rapport Médical", heading_style))
        # Découper le rapport en paragraphes
        for para in consultation.medical_report.split("\n"):
            if para.strip():
                story.append(Paragraph(para.strip(), body_style))
        story.append(Spacer(1, 0.3 * cm))

    # ── Pied de page ──────────────────────────────────────────────────────
    story.append(Spacer(1, 1 * cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#dee2e6")))
    story.append(Spacer(1, 0.2 * cm))
    story.append(Paragraph(
        f"Document généré le {datetime.utcnow().strftime('%d/%m/%Y à %H:%M')} UTC — "
        f"Healthcare AI Platform",
        label_style,
    ))

    doc.build(story)
    buffer.seek(0)
    return buffer


# ─────────────────────────────────────────────────────────────────────────────
#  POST /api/export/query-results
# ─────────────────────────────────────────────────────────────────────────────

@export_bp.route("/query-results", methods=["POST"])
@token_required
def export_query_results():
    """
    Exporte les résultats d'une requête SQL en CSV ou Excel.

    Body JSON:
        sql    (str, requis) : Requête SQL SELECT à exécuter.
        format (str)         : 'csv' ou 'excel' (défaut 'csv').

    Protections :
        - Seules les requêtes SELECT sont autorisées
        - Blacklist SQL appliquée
        - Limite de 10 000 lignes

    Returns:
        200 : Fichier CSV ou Excel
        400 : SQL invalide ou non-SELECT.
        403 : Opération non autorisée.
        500 : Erreur d'exécution.
    """
    body   = request.get_json(silent=True) or {}
    sql    = body.get("sql", "").strip()
    fmt    = body.get("format", "csv").lower()

    if not sql:
        return error_response("Le champ 'sql' est requis", 400)

    if fmt not in ("csv", "excel"):
        return error_response("Format invalide. Valeurs acceptées : 'csv', 'excel'", 400)

    # ── Vérifications de sécurité ──────────────────────────────────────────
    if not _check_sql_blacklist(sql):
        return error_response("Opération SQL interdite", 400, error="SQL_BLACKLISTED")

    if not sql.upper().strip().startswith("SELECT"):
        return error_response(
            "Seules les requêtes SELECT sont autorisées pour l'export",
            400,
            error="NON_SELECT_QUERY",
        )

    # ── Exécution SQL ──────────────────────────────────────────────────────
    start = time.perf_counter()
    try:
        from sqlalchemy import text
        result_proxy = db.session.execute(text(sql))
        rows         = result_proxy.fetchmany(10000)   # Limite 10k lignes
        columns      = list(result_proxy.keys())
        data         = [dict(zip(columns, row)) for row in rows]
        row_count    = len(data)
        exec_time    = round((time.perf_counter() - start) * 1000, 1)

    except Exception as exc:
        logger.error("Erreur export SQL — %s", str(exc))
        return error_response(f"Erreur d'exécution SQL : {str(exc)[:300]}", 500)

    if not data:
        return error_response("La requête n'a retourné aucun résultat", 404)

    # ── Génération du fichier ──────────────────────────────────────────────
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    if fmt == "csv":
        return _export_csv(data, columns, timestamp, row_count)
    else:
        return _export_excel(data, columns, timestamp, row_count)


def _export_csv(data: list, columns: list, timestamp: str, row_count: int):
    """Génère un fichier CSV depuis les données."""
    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=columns,
        extrasaction="ignore",
        quoting=csv.QUOTE_ALL,
    )
    writer.writeheader()

    for row in data:
        # Sérialiser les types non-string
        serialized = {}
        for k, v in row.items():
            if isinstance(v, datetime):
                serialized[k] = v.isoformat()
            elif v is None:
                serialized[k] = ""
            else:
                serialized[k] = str(v)
        writer.writerow(serialized)

    output.seek(0)
    csv_bytes = output.getvalue().encode("utf-8-sig")  # BOM pour Excel
    buffer    = io.BytesIO(csv_bytes)

    logger.info(
        "Export CSV — %d lignes, %d colonnes par user_id=%s",
        row_count, len(columns), "unknown",
    )

    return send_file(
        buffer,
        mimetype="text/csv; charset=utf-8",
        as_attachment=True,
        download_name=f"export_{timestamp}.csv",
    )


def _export_excel(data: list, columns: list, timestamp: str, row_count: int):
    """Génère un fichier Excel (.xlsx) depuis les données via openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fallback vers CSV si openpyxl n'est pas installé
        logger.warning("openpyxl non installé — fallback vers CSV")
        return _export_csv(data, columns, timestamp, row_count)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Résultats"

    # Style en-tête
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    # Écrire les en-têtes
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align

    # Écrire les données
    for row_idx, row in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row.get(col_name)
            if isinstance(value, datetime):
                value = value.isoformat()
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_len = max(
            len(str(cell.value or "")) for cell in col
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    logger.info("Export Excel — %d lignes, %d colonnes", row_count, len(columns))

    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"export_{timestamp}.xlsx",
    )
